import pandas as pd
from openpyxl import load_workbook

# Paths to the CSV and Excel files
csv_path = '../csv/02_xpaths.csv'
excel_file = '../excel/locators.xlsx'

# Read the CSV file into a DataFrame
csv_data = pd.read_csv(csv_path)

# Load the existing workbook
wb = load_workbook(excel_file)

# Create a new sheet in the workbook
new_sheet_name = 'test_test_2'
wb.create_sheet(new_sheet_name)

# Save the workbook to ensure the new sheet is created
wb.save(excel_file)

# Open the workbook with pandas ExcelWriter in append mode
with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
    # Write the CSV data to the new sheet
    csv_data.to_excel(writer, sheet_name=new_sheet_name, index=False)

print("Success")
